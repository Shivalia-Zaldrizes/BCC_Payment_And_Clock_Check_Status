import os
import csv
from dataclasses import dataclass
from typing import Dict, List, Optional
from openpyxl import load_workbook
from datetime import datetime
from app.logging import logger


# -----------------------------
# Normalized output structure
# -----------------------------

@dataclass
class PaymentUpdate:
    """
    One normalized payment fact.

    employee      -> must match Excel column header exactly
    week_ending   -> ISO date string: YYYY-MM-DD
    paid          -> True if paid, False otherwise
    source        -> CSV / Excel filename (for logging/debugging)
    """
    employee: str
    week_ending: str
    paid: bool
    source: str


# -----------------------------
# Helpers
# -----------------------------

def normalize_week_ending(value) -> Optional[str]:
    """
    Converts various date formats to YYYY-MM-DD.
    Returns None if invalid.
    """
    if not value:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt).date().isoformat()
        except ValueError:
            continue

    return None


def normalize_employee(name: str) -> Optional[str]:
    if not name:
        return None
    return str(name).strip()


def balance_is_paid(balance) -> bool:
    """
    Determines paid status.
    Paid = balance == 0
    """
    try:
        cleaned = str(balance).replace("$", "").replace(",", "").strip()
        return float(cleaned) == 0
    except Exception:
        return False


# -----------------------------
# QuickBooks CSV import
# -----------------------------

def parse_qb_csv(csv_path: str) -> List[PaymentUpdate]:
    """
    Expected columns (flexible names allowed):
      - Employee / Vendor / Name
      - Balance
      - Week Ending / Invoice Date (optional but recommended)
    """

    updates: List[PaymentUpdate] = []

    if not os.path.exists(csv_path):
        logger.warning(f"CSV not found: {csv_path}")
        return updates

    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            employee = normalize_employee(
                row.get("Employee")
                or row.get("Vendor")
                or row.get("Name")
            )

            if not employee:
                continue

            week_ending = normalize_week_ending(
                row.get("Week Ending")
                or row.get("Invoice Date")
            )

            if not week_ending:
                logger.warning(
                    f"Missing week ending in {csv_path} for {employee}"
                )
                continue

            paid = balance_is_paid(row.get("Balance"))

            updates.append(
                PaymentUpdate(
                    employee=employee,
                    week_ending=week_ending,
                    paid=paid,
                    source=os.path.basename(csv_path),
                )
            )

    logger.info(f"Parsed {len(updates)} payments from {csv_path}")
    return updates


def parse_csv_folder(folder_path: str) -> List[PaymentUpdate]:
    all_updates: List[PaymentUpdate] = []

    if not os.path.isdir(folder_path):
        logger.warning(f"CSV folder not found: {folder_path}")
        return all_updates

    for filename in os.listdir(folder_path):
        if filename.lower().endswith(".csv"):
            all_updates.extend(
                parse_qb_csv(os.path.join(folder_path, filename))
            )

    return all_updates


# -----------------------------
# Excel invoice source import
# -----------------------------

def parse_excel_invoice_file(
    excel_path: str,
    employee_col: str = "Employee",
    balance_col: str = "Balance",
    week_col: str = "Week Ending",
) -> List[PaymentUpdate]:
    """
    Reads an Excel file that lists invoices/payments per employee.
    """

    updates: List[PaymentUpdate] = []

    if not os.path.exists(excel_path):
        logger.warning(f"Excel source not found: {excel_path}")
        return updates

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

    if (
        employee_col not in headers
        or balance_col not in headers
        or week_col not in headers
    ):
        logger.error(
            f"Missing required columns in {excel_path}: "
            f"{employee_col}, {balance_col}, {week_col}"
        )
        return updates

    for row in ws.iter_rows(min_row=2):
        employee = normalize_employee(row[headers[employee_col]].value)
        if not employee:
            continue

        week_ending = normalize_week_ending(
            row[headers[week_col]].value
        )
        if not week_ending:
            continue

        paid = balance_is_paid(row[headers[balance_col]].value)

        updates.append(
            PaymentUpdate(
                employee=employee,
                week_ending=week_ending,
                paid=paid,
                source=os.path.basename(excel_path),
            )
        )

    logger.info(f"Parsed {len(updates)} payments from {excel_path}")
    return updates


def parse_excel_folder(folder_path: str) -> List[PaymentUpdate]:
    all_updates: List[PaymentUpdate] = []

    if not os.path.isdir(folder_path):
        logger.warning(f"Excel folder not found: {folder_path}")
        return all_updates

    for filename in os.listdir(folder_path):
        if filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            all_updates.extend(
                parse_excel_invoice_file(
                    os.path.join(folder_path, filename)
                )
            )

    return all_updates


# -----------------------------
# Master collector
# -----------------------------

def collect_all_payments(
    csv_folder: Optional[str] = None,
    excel_folder: Optional[str] = None,
) -> List[PaymentUpdate]:
    """
    Single entry point used by main.py
    """

    updates: List[PaymentUpdate] = []

    if csv_folder:
        updates.extend(parse_csv_folder(csv_folder))

    if excel_folder:
        updates.extend(parse_excel_folder(excel_folder))

    # Deduplicate: last source wins
    deduped: Dict[tuple, PaymentUpdate] = {}
    for u in updates:
        key = (u.employee, u.week_ending)
        deduped[key] = u

    final_updates = list(deduped.values())

    logger.info(
        f"Collected {len(final_updates)} unique payment updates"
    )

    return final_updates
